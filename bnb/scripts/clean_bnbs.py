'''Very ugly and lengthy code to complete and finish all models.
Command needed: python manage.py runscript load_data
Please don't read or use xx
Layla Hoogeveen, Information Organization, October 2021
'''

import openpyxl
import requests
from pathlib import Path
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from bnbs.models import Accommodation
import os
import sys
import re


def run():
    '''Remove accommodations with invalid urls by scraping'''

    xlsx_file = Path('data\curated_data\clean', 'airbnb_data_final.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active

    total = 0
    i = 0
    options = Options()
    options.headless = True
    options.add_argument("--log-level=3")
    row_count_before = sheet.max_row
    print ("Rows before:", row_count_before)

    driver = webdriver.Chrome(executable_path=r'C:/Users/Layla/.wdm/drivers/chromedriver/win32/94.0.4606.61/chromedriver.exe', options=options)

    for rowNum in range(14198, 1, -1):
        
        total = total + 1
        URL = sheet.cell(row = rowNum, column = 19).value
        
        if check_url(URL, driver) == True:
            i=i+1
            print ("\n", i, "van de", total, "\n\n")
            sheet.delete_rows(rowNum)
        
        if total % 50 == 0:
            print ("\n", i, "van de", total, "\n\n")
        if total % 10 == 0:
            wb_obj.save(xlsx_file)
        if total / 500 == 1:
            break

    wb_obj.save(xlsx_file)
    row_count_after = sheet.max_row

    print ("\n\n", "Final", "\n")
    print (i, "van de", total, "verwijderd", "\n")
    print ("Oorspronkelijk aantal accommodaties: 16478")
    print ("Accommodations before running script:", row_count_before, "\n")
    print ("Accommodations after running script:", row_count_after)

    driver.quit()



def check_url(url, driver):
    '''Check whether url is invalid by searching for error message'''

    page = requests.get(url)
    driver.get(url)
    time.sleep(5)
    res=driver.page_source
    soup = BeautifulSoup(res, "html.parser")

    # Not the fastest, but the only way to find this specific text
    soup = str(soup)
    if "Permission denied by Himeji" in soup:
        return True
    
    return False


# def run():
#     '''Remove accommodations with invalid urls by scraping'''

#     num_accs_before = len(Accommodation.objects.all())
#     print (num_accs_before)
#     accs = Accommodation.objects.filter(pk__gt=0).all()

#     total = 0
#     i = 0
#     options = Options()
#     options.headless = True
#     options.add_argument("--log-level=3")

#     driver = webdriver.Chrome(executable_path=r'C:/Users/Layla/.wdm/drivers/chromedriver/win32/94.0.4606.61/chromedriver.exe', options=options)
#     for a in accs:
#         total = total + 1

        
#         URL = a.listing_url
#         if check_url(URL, driver) == True:
#             i=i+1
#             print ("\n", i, "van de", total, "\n\n")
#             a.delete()
#         if total > 200:
#             break
#         if total % 50 == 0:
#             print (total)
        

#     print ("\n\n", "Final", "\n")
#     print (i, "van de", total, "verwijderd", "\n")
#     print ("Accommodations before running script:", num_accs_before, "\n")
#     print ("Accommodations after running script:", len(Accommodation.objects.all()))
#     driver.quit()
